import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from universities.models import University, UniversityIndicator
from universities.services.university_normalizer import (
    canonical_university_name,
    normalize_university_name,
)


SOURCE_URL = "https://www.academyinfo.go.kr/index.do"

# alias, 저장 단위, 원본 값 배율
INDICATOR_COLUMNS = [
    ("EMPLOYMENT_RATE", ["취업률"], "%", Decimal("1")),
    ("AVG_TUITION", ["평균등록금", "연평균등록금"], "KRW", Decimal("1000")),
    (
        "SCHOLARSHIP_PER_STUDENT",
        ["학생1인당연간장학금", "학생1인당장학금"],
        "KRW",
        Decimal("1"),
    ),
    (
        "DORMITORY_CAPACITY_RATE",
        ["기숙사수용률", "기숙사수용율"],
        "%",
        Decimal("1"),
    ),
    (
        "EDUCATION_COST_PER_STUDENT",
        ["학생1인당교육비"],
        "KRW",
        Decimal("1000"),
    ),
    (
        "STUDENTS_PER_FULLTIME_FACULTY",
        ["전임교원1인당학생수"],
        "PERSON",
        Decimal("1"),
    ),
    (
        "FULLTIME_FACULTY_SECURING_RATE",
        ["전임교원확보율학생정원기준", "학생정원기준전임교원확보율"],
        "%",
        Decimal("1"),
    ),
]

# 대학알리미의 본분교 표기를 K-unirank의 대학 단위 페이지 정책에 맞춘다.
CAMPUS_TARGETS = {
    ("건양대학교", "제2캠퍼스"): "건양대학교 메디컬캠퍼스",
    ("경기대학교", "제2캠퍼스"): "경기대학교 서울캠퍼스",
    ("경동대학교", "제3캠퍼스"): "경동대학교 메디컬캠퍼스",
    ("경동대학교", "제4캠퍼스"): "경동대학교 메트로폴캠퍼스",
    ("단국대학교", "본교"): "단국대학교 죽전캠퍼스",
    ("단국대학교", "제2캠퍼스"): "단국대학교 천안캠퍼스",
    ("명지대학교", "본교"): "명지대학교 자연캠퍼스",
    ("명지대학교", "제2캠퍼스"): "명지대학교 인문캠퍼스",
    ("상명대학교", "제2캠퍼스"): "상명대학교 천안캠퍼스",
    ("신한대학교", "제2캠퍼스"): "신한대학교 동두천캠퍼스",
    ("안양대학교", "제2캠퍼스"): "안양대학교 강화캠퍼스",
    ("예원예술대학교", "본교"): "예원예술대학교 전북희망캠퍼스",
    ("예원예술대학교", "제2캠퍼스"): "예원예술대학교",
    ("을지대학교", "본교"): "을지대학교 대전캠퍼스",
    ("을지대학교", "제2캠퍼스"): "을지대학교",
    ("을지대학교", "제3캠퍼스"): "을지대학교 의정부캠퍼스",
    ("인천가톨릭대학교", "본교"): "인천가톨릭대학교 강화캠퍼스",
    ("인천가톨릭대학교", "제2캠퍼스"): "인천가톨릭대학교",
    ("전남대학교", "제2캠퍼스"): "전남대학교 여수캠퍼스",
    ("중앙대학교", "제2캠퍼스"): "중앙대학교 다빈치캠퍼스",
    ("홍익대학교", "제2캠퍼스"): "홍익대학교 세종캠퍼스",
}

SOURCE_NAME_TARGETS = {
    "가야대학교(김해)": "가야대학교",
    "건국대학교(글로컬)": "건국대학교 글로컬캠퍼스",
    "고려대학교(세종)": "고려대학교 세종캠퍼스",
    "동국대학교(WISE)": "동국대학교 WISE캠퍼스",
    "연세대학교(미래)": "연세대학교 미래캠퍼스",
    "영산대학교(양산)": "영산대학교 양산캠퍼스",
    "영산대학교(해운대)": "영산대학교",
    "한양대학교(ERICA)": "한양대학교 ERICA캠퍼스",
}

# 이미 현재 대학으로 통합돼 있고, 같은 엑셀에 현재 대학 행이 따로 존재하는 구교명.
SKIP_SOURCE_NAMES = {
    "가야대학교(고령)",
    "경남도립거창대학",
    "경남도립남해대학",
    "경북도립대학교",
    "국립강릉원주대학교",
    "국립목포대학교(담양캠퍼스)",
    "국립창원대학교(거창캠퍼스)",
    "국립창원대학교(남해캠퍼스)",
}


def normalize_header(value):
    text = str(value or "").strip()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[()\[\]{}·,:/_-]", "", text)
    return text


def clean_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_decimal(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {"-", "--", "N/A", "n/a"}:
        return None
    text = text.replace(",", "").replace("%", "").replace("원", "").strip()
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_year(value):
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def year_from_header(header):
    match = re.search(r"20\d{2}", str(header or ""))
    return int(match.group(0)) if match else None


def resolve_target_name(raw_name, campus_name):
    raw_name = clean_text(raw_name)
    campus_name = clean_text(campus_name)

    if not raw_name or raw_name in SKIP_SOURCE_NAMES:
        return None, 0

    # 과거 산업대 행은 현재 대학 행과 중복되는 0값 레거시 행이므로 사용하지 않는다.
    if "(산업대)" in raw_name:
        return None, 0

    # K-unirank는 한국폴리텍 캠퍼스를 대학 하나로 접어 표시한다.
    # 임의의 한 캠퍼스 수치로 대표값을 만들지 않기 위해 이번 대학 단위 지표에서는 제외한다.
    if raw_name.startswith("한국폴리텍"):
        return None, 0

    direct = SOURCE_NAME_TARGETS.get(raw_name)
    if direct:
        return direct, 100

    campus_target = CAMPUS_TARGETS.get((raw_name, campus_name))
    if campus_target:
        return campus_target, 100

    # 가톨릭대학교는 서비스에서 캠퍼스를 한 대학으로 접으므로 본교 수치만 대표로 사용한다.
    if raw_name == "가톨릭대학교" and campus_name not in {"", "본교"}:
        return None, 0

    # 강원대는 통합 대학 단위로 관리한다. 본교 외 행이 대표값을 덮지 않게 한다.
    if raw_name == "강원대학교" and campus_name not in {"", "본교"}:
        return None, 0

    canonical = canonical_university_name(raw_name)
    return canonical or raw_name, 50 if campus_name in {"", "본교"} else 20


def should_skip_zero(code, value, graduate_count, student_count, admission_quota):
    if value != 0:
        return False
    if code == "EMPLOYMENT_RATE":
        return not graduate_count
    if code in {
        "STUDENTS_PER_FULLTIME_FACULTY",
        "FULLTIME_FACULTY_SECURING_RATE",
        "SCHOLARSHIP_PER_STUDENT",
        "EDUCATION_COST_PER_STUDENT",
        "DORMITORY_CAPACITY_RATE",
    }:
        return not student_count
    if code == "AVG_TUITION":
        return not student_count and not admission_quota
    return False


class Command(BaseCommand):
    help = (
        "대학알리미 '대학주요정보 한눈에 보기' XLSX에서 K-unirank 핵심 공식 지표를 "
        "자동 탐지해 일괄 적재합니다. 기본은 미리보기이며 --apply 시 반영합니다."
    )

    def add_arguments(self, parser):
        parser.add_argument("path", help="대학알리미에서 내려받은 XLSX 파일")
        parser.add_argument("--sheet", help="시트명. 생략하면 첫 시트를 사용합니다.")
        parser.add_argument(
            "--header-row",
            type=int,
            default=1,
            help="헤더 행 번호 (1부터 시작, 기본 1)",
        )
        parser.add_argument(
            "--year",
            type=int,
            help="모든 지표의 공시연도를 파일 헤더 대신 강제로 지정합니다.",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="실제 DB에 반영합니다.",
        )

    def handle(self, *args, **options):
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"파일을 찾을 수 없습니다: {path}")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise CommandError("이 명령은 대학알리미 XLSX/XLSM 파일을 대상으로 합니다.")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl 설치가 필요합니다.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        if options.get("sheet"):
            sheet_name = options["sheet"]
            if sheet_name not in workbook.sheetnames:
                raise CommandError(
                    f"시트를 찾을 수 없습니다: {sheet_name} / "
                    f"가능한 시트: {', '.join(workbook.sheetnames)}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        header_row = max(options["header_row"], 1)
        raw_headers = [cell.value for cell in worksheet[header_row]]
        normalized_headers = [normalize_header(value) for value in raw_headers]

        university_col = self.find_column(normalized_headers, ["학교명", "대학명"])
        campus_col = self.find_column(normalized_headers, ["본분교명", "캠퍼스명"])
        graduate_col = self.find_column(normalized_headers, ["졸업생수"])
        student_col = self.find_column(normalized_headers, ["재학생"])
        admission_col = self.find_column(normalized_headers, ["입학정원"])
        if university_col is None:
            raise CommandError("학교명/대학명 열을 찾지 못했습니다.")

        detected = {}
        for code, aliases, unit, multiplier in INDICATOR_COLUMNS:
            column_index = self.find_column(normalized_headers, aliases)
            if column_index is not None:
                detected[code] = {
                    "column": column_index,
                    "unit": unit,
                    "multiplier": multiplier,
                    "header": raw_headers[column_index],
                    "header_year": year_from_header(raw_headers[column_index]),
                }

        if not detected:
            raise CommandError(
                "핵심 지표 열을 하나도 찾지 못했습니다. --header-row 값을 확인하세요."
            )

        self.stdout.write("[자동 탐지된 지표]")
        for code, info in detected.items():
            self.stdout.write(f"- {code}: {info['header']}")

        universities = list(University.objects.filter(is_active=True))
        university_map = {
            normalize_university_name(university.name): university
            for university in universities
        }

        # 같은 대학/연도/지표 후보가 여러 캠퍼스에서 들어오면 우선순위가 높은 행만 유지한다.
        prepared = {}
        unmatched = set()
        rows_seen = 0

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            rows_seen += 1
            raw_name = row[university_col] if university_col < len(row) else None
            campus_name = (
                row[campus_col]
                if campus_col is not None and campus_col < len(row)
                else ""
            )
            target_name, priority = resolve_target_name(raw_name, campus_name)
            if not target_name:
                continue

            university = university_map.get(normalize_university_name(target_name))
            if university is None:
                if clean_text(raw_name):
                    unmatched.add(f"{clean_text(raw_name)} [{clean_text(campus_name) or '본교'}]")
                continue

            graduate_count = self.numeric_cell(row, graduate_col)
            student_count = self.numeric_cell(row, student_col)
            admission_quota = self.numeric_cell(row, admission_col)

            for code, info in detected.items():
                column_index = info["column"]
                if column_index >= len(row):
                    continue
                value = parse_decimal(row[column_index])
                if value is None:
                    continue
                if should_skip_zero(
                    code,
                    value,
                    graduate_count,
                    student_count,
                    admission_quota,
                ):
                    continue

                indicator_year = options.get("year") or info["header_year"]
                if indicator_year is None:
                    continue

                value *= info["multiplier"]
                key = (university.pk, indicator_year, code)
                candidate = {
                    "university": university,
                    "year": indicator_year,
                    "indicator_code": code,
                    "value": value,
                    "unit": info["unit"],
                    "source_label": f"대학알리미 · {clean_text(campus_name) or '본교'} · {info['header']}",
                    "priority": priority,
                }
                previous = prepared.get(key)
                if previous is None or priority > previous["priority"]:
                    prepared[key] = candidate

        items = list(prepared.values())
        self.stdout.write(
            f"읽음 {rows_seen}행 / 적재 후보 {len(items)}건 / 미매칭 대학 {len(unmatched)}개"
        )
        if unmatched:
            self.stdout.write(self.style.WARNING("[미매칭 대학 예시]"))
            for name in sorted(unmatched)[:40]:
                self.stdout.write(f"- {name}")

        if not options["apply"]:
            self.stdout.write(
                self.style.WARNING("미리보기 완료. 실제 반영은 --apply를 추가하세요.")
            )
            return

        created = 0
        updated = 0
        with transaction.atomic():
            for item in items:
                _, was_created = UniversityIndicator.objects.update_or_create(
                    university=item["university"],
                    year=item["year"],
                    indicator_code=item["indicator_code"],
                    source="ACADEMYINFO",
                    defaults={
                        "value": item["value"],
                        "unit": item["unit"],
                        "source_url": SOURCE_URL,
                        "source_label": item["source_label"],
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(
            self.style.SUCCESS(f"반영 완료: 신규 {created}건 / 갱신 {updated}건")
        )

    @staticmethod
    def find_column(headers, aliases):
        normalized_aliases = [normalize_header(alias) for alias in aliases]
        for index, header in enumerate(headers):
            for alias in normalized_aliases:
                if alias and alias in header:
                    return index
        return None

    @staticmethod
    def numeric_cell(row, index):
        if index is None or index >= len(row):
            return Decimal("0")
        return parse_decimal(row[index]) or Decimal("0")
