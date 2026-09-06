import csv
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from universities.models import University, UniversityIndicator
from universities.services.indicators import CORE_INDICATOR_CODES


DEFAULT_UNIVERSITY_COLUMN = "학교명"
DEFAULT_YEAR_COLUMN = "기준연도"


def normalize_university_name(value):
    return re.sub(r"\s+", "", str(value or "").strip())


def parse_year(value):
    if value is None or value == "":
        raise ValueError("empty year")
    return int(float(str(value).strip()))


def parse_decimal(value):
    if value is None:
        raise ValueError("empty value")
    text = str(value).strip()
    if not text or text in {"-", "--", "N/A", "n/a"}:
        raise ValueError("empty value")
    text = text.replace(",", "").replace("%", "").replace("원", "").strip()
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"invalid numeric value: {value}") from exc


class Command(BaseCommand):
    help = (
        "대학알리미/공공데이터 공식 CSV 또는 XLSX 한 파일에서 대학 단위 지표 하나를 "
        "UniversityIndicator로 가져옵니다. 기본은 미리보기이며 --apply 시 DB에 반영합니다."
    )

    def add_arguments(self, parser):
        parser.add_argument("path", help="CSV/XLSX 파일 경로")
        parser.add_argument(
            "--indicator",
            required=True,
            choices=CORE_INDICATOR_CODES,
            help="저장할 K-unirank 표준 지표 코드",
        )
        parser.add_argument(
            "--value-column",
            required=True,
            help="공식 파일에서 값이 들어있는 열 이름",
        )
        parser.add_argument(
            "--university-column",
            default=DEFAULT_UNIVERSITY_COLUMN,
            help=f"대학명 열 이름 (기본: {DEFAULT_UNIVERSITY_COLUMN})",
        )
        parser.add_argument(
            "--year-column",
            default=DEFAULT_YEAR_COLUMN,
            help=f"기준연도 열 이름 (기본: {DEFAULT_YEAR_COLUMN})",
        )
        parser.add_argument("--unit", default="", help="저장할 원본 단위")
        parser.add_argument(
            "--source-url",
            default="https://www.academyinfo.go.kr/index.do",
            help="원문/공식 데이터 출처 URL",
        )
        parser.add_argument(
            "--source-label",
            default="대학알리미 대학정보공시",
            help="출처 설명",
        )
        parser.add_argument(
            "--sheet",
            help="XLSX 시트명. 생략하면 첫 번째 시트를 사용합니다.",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="실제 DB에 저장합니다. 없으면 미리보기만 합니다.",
        )

    def handle(self, *args, **options):
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"파일을 찾을 수 없습니다: {path}")

        rows = list(self.read_rows(path, options.get("sheet")))
        if not rows:
            raise CommandError("읽을 데이터가 없습니다.")

        headers = set(rows[0].keys())
        required = {
            options["university_column"],
            options["year_column"],
            options["value_column"],
        }
        missing = sorted(required - headers)
        if missing:
            raise CommandError(
                "필수 열을 찾지 못했습니다: "
                + ", ".join(missing)
                + "\n현재 열: "
                + ", ".join(str(item) for item in rows[0].keys())
            )

        universities = list(University.objects.filter(is_active=True))
        university_map = {
            normalize_university_name(university.name): university
            for university in universities
        }

        prepared = []
        unmatched = set()
        invalid_rows = []

        for row_number, row in enumerate(rows, start=2):
            raw_name = row.get(options["university_column"])
            normalized_name = normalize_university_name(raw_name)
            university = university_map.get(normalized_name)
            if university is None:
                if str(raw_name or "").strip():
                    unmatched.add(str(raw_name).strip())
                continue

            try:
                year = parse_year(row.get(options["year_column"]))
                value = parse_decimal(row.get(options["value_column"]))
            except ValueError as exc:
                invalid_rows.append((row_number, str(raw_name or ""), str(exc)))
                continue

            prepared.append((university, year, value))

        self.stdout.write(
            f"읽음 {len(rows)}행 / 매칭 {len(prepared)}행 / "
            f"미매칭 대학 {len(unmatched)}개 / 값 오류 {len(invalid_rows)}행"
        )

        if unmatched:
            self.stdout.write(self.style.WARNING("[미매칭 대학 예시]"))
            for name in sorted(unmatched)[:30]:
                self.stdout.write(f"- {name}")

        if invalid_rows:
            self.stdout.write(self.style.WARNING("[값 오류 예시]"))
            for row_number, name, reason in invalid_rows[:20]:
                self.stdout.write(f"- {row_number}행 {name}: {reason}")

        if not options["apply"]:
            self.stdout.write(
                self.style.WARNING("미리보기만 완료했습니다. 실제 반영은 --apply를 추가하세요.")
            )
            return

        created = 0
        updated = 0
        with transaction.atomic():
            for university, year, value in prepared:
                _, was_created = UniversityIndicator.objects.update_or_create(
                    university=university,
                    year=year,
                    indicator_code=options["indicator"],
                    source="ACADEMYINFO",
                    defaults={
                        "value": value,
                        "unit": options["unit"],
                        "source_url": options["source_url"],
                        "source_label": options["source_label"],
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(
            self.style.SUCCESS(f"반영 완료: 신규 {created}건 / 갱신 {updated}건")
        )

    def read_rows(self, path, sheet_name=None):
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                yield from csv.DictReader(handle)
            return

        if suffix in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise CommandError(
                    "XLSX를 읽으려면 openpyxl이 필요합니다. requirements 설치 후 다시 실행하세요."
                ) from exc

            workbook = load_workbook(path, read_only=True, data_only=True)
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise CommandError(
                        f"시트를 찾을 수 없습니다: {sheet_name} / "
                        f"가능한 시트: {', '.join(workbook.sheetnames)}"
                    )
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook[workbook.sheetnames[0]]

            values = worksheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip() for value in next(values)]
            except StopIteration:
                return

            for values_row in values:
                yield dict(zip(headers, values_row))
            return

        raise CommandError("지원 파일 형식은 .csv, .xlsx, .xlsm 입니다.")
